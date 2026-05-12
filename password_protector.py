#!/usr/bin/env python3
"""
LOA PDF Generator v2  –  Change-detection edition
Run:  python3 loa_generator_v2.py
Opens automatically at http://127.0.0.1:5051

Requirements:  pandas  openpyxl  reportlab
Install:  pip install pandas openpyxl reportlab
"""

import os, io, json, zipfile, threading, webbrowser, base64, tempfile, math, re, secrets
from http.server import HTTPServer, BaseHTTPRequestHandler
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as rl_canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAHYAAAAnCAIAAABVFysUAAAQOklEQVR4nO2ae3BUZZbAz/luP9Kd7jxJP/J+dSJCSAhDEoIyPOSliK+CMDg7WrWLQ5xl/9gtq3RWXctaZ0qdLa0dHV4jjro4A1FA0IEACajACqxJQJIgSac7SSdp0p10mk76ee939o8OIaY7EWZBXYvzR1f1vef7vnN/99zz+O5FIoLbciuFfd8G/PjlNuJbLrcR33KJgpiIJElyuVy3w/RNEVnkIc55S0uLw+FIS0srKChAxO/erB+TTPRiSZJaW1sdDgci9vT0tLW13fbl/6Nc82Ii4py3trb29/eHPRcRbTYbAJhMpuv0ZQIAIkAAACSA20/AeMSc84sXL4b9d+xgmDIi5ufnXw9lLoZCHg8AMrlMHhuL8N0h5px7vV6FQqFQKG5oYDAYDAaDsbGxNzEkElEgEJAkSa1WM7jqv319ff39/ZFhARF7e3ttNtv1RAyP1XpkzZqja9Y0v/EmfYd8AcDv9z/z62c+//zzGx14/NPjTz/ztChJN9eempqaV3/3OwjHYiLq6+sbHh42GAyRd5IxptFo2trazGbzt1JGAOAcOQei7zhGIOLfljYI4KbnmzDG8LSMc2632z0eT/iEXq8frxrm63a7AaCrq6ujo+PGrCEgIiKCKQYREBEPq/0Yhdnt9jDB0f+MjVEe44tXpbOz84Yoc+Ki3yf6vBLxKXSkYEAcGZaCgaln5pwHg0G/38/5pLNNNtDv9wcCgW8dSESiKPp8PlEUpzaGiEKhkM/nkyRpak3ZlStXJgSHMGWHwzHGd/zZzs5ORMzJyZk6ORCQo6Ghfed7QxcvEfEEk6nw7zckziwSBLx2MYGA7fDhviOH3RYLD4YEhVybn5e+fEXakqVMJmPs2vyiKJ46derEqZOX7XZOlJSYWDa3bPGSJW1tbUeOHHmyulqlUkWlMDQ0dPjw4XPnz3k8w8hQr9NVlFcsWLBALpdHKnd3dx+qrf360td+n0+hVObn5i1dujQyzxNRc3NzXX2d1doZEkNqtXrmnTOWL18eExPzxx1vrbpvVWFBwTcQR7VMEJiBYa/bPSFjhRezWq0AMDXlgbNne+vqgPNwfHZduHD66acqX3s98Y7pAMA5eR39Xz7/vKftEgACEAGQGHKdv+A6/1XP0fo5L7yg1GrCUw2PjGzZsqWjwzz9jull969WqVT9/ZdPnDrZ2NRYXFJibm+P6kREZDabt27bGgqF5s4tS09PD4VCF1tbP9yzp7GpcePGak1s7Hj9M2fP7Nq1q7Cg8KEHH9Rq45xOxxdfnH799dcfeOCBpUuXjl2pxPmePXuOHas3Go3Lli6NT0hwDw01NDW+8sora6uq2tvavCMjEyyJQEzEEIeP1/fVfDitao0vz0QIE2ovRLRarYiYnZ09GeXA0BDK5cZFi1mMwn78U3FkhAJBc01N6bPPMUAeCHz57L8Om80AyJTKpOJZap3e53QMNDTyYGDwXGPP0cO5Dz0MAJzzt9/eYbN1b/zlxqKiorE0snLFynfefaf2UO1k99jlcm3dujUuPr5648bk5OTwwCWLFzeea/rT23965513qjduZGy08+Kcanbvfuyxx+bMLh27ogV3L9izd8++ffuSkpLmzJkTTqdH647W19evWLnivnvvlQmj9JYtW/bxJ5+8//7OqJbIOOdjKxEQQ+b97FjfrhokcO7anbJu3UhuPgFNQImIFosFAKagXPLsc/rK+YiYUlbR+PxzgMxj7iDOOWL3wb96zGYAUKen/eSll2JT0wCAAL3dne1//nPOw4/E5ZvCKFtaWlpaWqrWrRvjG149Jibm8cce//eXXhpwOiOXJqJDtbXBUPCXTzwxbdq08WbPLi4ZeujBmt27W1tbZ8yYMXZqyT1Lx/MFAJlM9sjDj3R1de37aF9xSbFcJh8eGT7414PFJcX337dqDBoACIJw/6pVdrv9XFNTpDHs8uXL4SRAQAzQe+J4z192IyAgIgfHrr/EdrRHJRimbLVaoz6n2nyToXK+TGCMYUrxLAjHPlECBCKw1R1FQCIo+ud/0aZlCIIgCIJMYHHZOaXP/DqxsFAQRqvJhqZGjUZTWT4v8kYqFIqKioqotkmSdK6pqbS0NCUlJdLs+ZWVGo22obFxzHJEWHDXXVEL1oULFw44B7q6uoiotbU1EPAvveee8XzHNO9ZsiRq5cQAwG63c+IMme/kZz3v7/qGQRI4du1SWc2TlVQWi6WzszPypNpoCKcrBBBUKia7ll6IuLevDwBkmtiE6XfA5B0KIjodDqPBGJmdYLTE1EUdGPAFrniupGdkRD2rkCsMev3AgHPMbE2sRqvVRl0iPS0dGDidTgAYGBgQmJCelh51WqPRGPVaWHii/sv9zOW07dqNEVoocdfefYY4DUYwDldyVqtVimiNBIWCrj3XDIVrtx0BGBMAgDgHzmGqmhmUSqUoilFPEVEwGIw6GBkiIpekyYKYKElhG8ISEkOTFV6hUAgIBEEAAEGQEdBk9kxWEY5euTZO2+0LpKyrIjbRJlQqc3/1K6aJ0+l1Ue0oLCwMW/CNUWy0OiMAAMRx1wPI4vLzAIh7ff1nznAEClcUxDmRFAqNb0OysrJtPbZwZxQpFy9+HfUOxahi9Dp9W3s7j2bwiNfb29ubkZE+dgP8fr+1szNSkxO1XGxFxPT0DADIzMjgnF+6dClSk4g6OjqiGskQMT4+3j3kBoCR7NxpP1s3njIqFTn/tAmycgBAEAS9Xj+BcmFhYdS2ewpBhhn3riIAQLjw2uv2E59LwRBwCnl9nR8fqH90ve3oEc4JABCxct48Jgg1H9ZEPijt5vaGhoaoO02IWF5R3nzhgrl9YiIhooOHDkqSWF5Wfi1/Ah44sD/SPV1ud93RunyTSZeSgoh5+fkpKSkff/Kx3++foBkKhQ7VHopKgcXFxY32b4AA6M3J0131ZVTIczZtguy8se0GQRB0umu+XFhYaDQab3SDChH1lZWGhYsAQLzibvy3F+qqquof/0Vd1drm114LOpznf/vbvlMnwsqJiYlVa9c2NDRs3rrFYrWKoiiJksftqa+v/8PmzWkZ6RSt80bExYsWZ6RnbNu+/fxX5zkPPxgUCAY+OrD/WH398uXLjUbjuAHgcg1t++N2t9tNV8Vqtb7x+/8URXFtVVX4GmWCsH79o/39/Zu3bnE6naN6nFwDrq3btrlcrqjpTja+e0ZEAhrOy9dVrXHu/SjnH5+EnFxAGu8pMplMp9M5HI6CgoJIvkSAo8sQESEiEBHw0c0WICAAAhSE4qeeElQxPbW1wHnIPRRyuwGIEJhckf+LvzNW3jVmUkV5hVarrfngg1deeVmpjFHI5cMjI9o4bdXaKn/Ab+vsigxTACCXy6urq3fs2LF58xaj0WgwGsSQaLFYvF7v8mXL77v3vgl7ttXV1W/teOu555/PyclRq9WDgwPd3d1JScmbNm1KNRjHNtALCwo2bHji3ffeffHFF00FBQkJ8S7XUHt7W2ZG5oZ/2PDqf7wa5X7X19dPOEREgJCqiQVNPH2T75jodLrExMRI//W7XLYjhxFBk5mtKytDRCCQuNR5YD8PBWWxmowVKxkyROBERNzTYbF/dtxjsYhev1yriTcVpC5ZHKMzMCTE0botPLMkSVar1dbTQ8QNekNeXp5cLt+/f//Jkyd/85vfyGSyUCj06WefmvJNWVlZY/ZIktTc3PzVha8GBwcZE1JTjWVzy1JTU8db3mGxdHSYFy1axCX+ZUNDc0vzsMej0WhMBQVlc8tilMpIPh6P5/SZ0xaLxefzJSYmFc2cWVRUFAqFTpw8UVQ0y6DXE1FzS4vbPVQ5rxLr6uome9K/8SiNE71en5CQcLM2sDkfzUkEIDAcPy0RNZ07d/r0Fw8+8KBer5+woiiJL7/8cnLytCc2bIgsVMdPMv7vt2ytXLfy9WuyrKysyeqVvr6+yLl0Ot1N5AsAjDFBYILAZAKL7CELTKbubtuWrVv7+vrGLCEiv9//Xzt39vb1RW0EJkwyXqY25vqVb0BTkqRw+zCFL49G1cnjw60TIurp7d2+fdvAwKDJZEpPT5PJZIODgy2trcFAYN26n1WUlyMi5+Jwr537/UwVI9doZTEqQSbzDbliEhOkQECuVAFicHgYZULQc0V0e2QatSpFxwQBETmn0LBnoKkpeXYpU8iHu7oBUaaOUcTHe/suM7kQm5oW9HiUGg3K5UA0Mjgol8lkKjXK5YEBp6ulWRGfkDRzBgF4Oq1ATFDKYw1GR8OX6tTU2NQ0DL9VmoIyIhoMhu+F75gEgsGmpsYLX11wOJ2iKMZptXn5eeXlFclJSYhIAFySmt94M7mkWJmQ4LXbE2fMUKca23fuzFu3/sKbvzetrVLpjL0nP9OkpvZ+ejyhoCDgvuK2dNzx2OPKuDgiaNm+NfXuBW6LJaV0tmX3B0mzS2KSEqVAYLClJdZgdFstaXcvGO7tSV20WPL5Oj6oiUmellxSMmKzDTY3py642zcwMNTenrt6dcsfNuvnzZNr4/yuQUTwOZ3Zqx+QAQBjLCcnh4i6uroiCRKR3W6fNWvW98UXAJQKRXlZedncMuCAiHS10RxN9ACIwOQyRUKCTKO5mp5HbVXGx1s/+ST34UcAARCYUmm8+24OGBoY7KqtNa1ZA0Ca9IwRuz1zxQp/f7+gVisTEuSxWh70x+fnxeXmuVpb40ym3s9PpC6QeuvrUxcuHLzQjESOs2fvfLKaIdPm5et+UiZ5RwS1WpGQKI9VyeM0tmPHC37+cyaTjUYxxlhubm5mZmZkXCai3Nzc75HvmCAiCggMokZAIuKhEBeDY18WECAgMibLf3S9+YPdktd3VRUYoCIhnoIBAGCImStXMplg2buXCIiLUjAgiSEAtB448PV77+WvX88YSy6e5Wxo8Nrt6rTwviAwhQKAISADZAwBgUuSFPSLoZA2Oydz+bKWzZtBkq4lijDljIyM8ZSJKC8vLyMj43vn+62i0GqnlZYmFtwh18Z5+/uJS6LfH/ZouUpVsG59z7F6IEDikiiFRkbMe/ckFc0CAEmS7P/9hb5y/oitG5E0GVnJxSVx2TkAkLe2Sq3XSwE/AKTMLrXs22eYN48hAwBAlMWqhlqbRS6F/H7b8eNEEJtqnDarJD4vr//s/8To9LFGo7f/8sS3tpxzs9nc3d2NiGG+mZmZP3S+BBLnjsbGlDmzGSIXJcuePQHPFX3FvKTp0/sbGvSlcwDB53QQMndrq7vDwgRh2uySxDunM2AS0eVTp1wXW5JLShNM+eYPP2SCEJuenlBQiAxjpqVcPnPaOH8+Q3R3dGizcxhD59cXNYY0hVrVWXsw4BgAxtIW/jTGYGh/fyfKZDEpKUkzZ/YcPqxKSs5avTrKi3HOeXt7u81my83NzcrK+qHzBYBw40iEVxuW8BYHICEiSQQMGSJx4gAIBAhAAAhstLsBGn17G0aBV38JAAEpPAIROUF4n4ETIbDRfhUQIPzOgka/hLpGFBmL9u1B+IjP51OpVP8v+P7A5W/8vOO2XL/c/oT7lsttxLdcbiO+5fK/oHMo4ZdPlUYAAAAASUVORK5CYII="

# ── HTML ──────────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>LOA Generator · HC Global Fund Services</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --red:#c0392b;--red-dark:#a93226;--red-ring:rgba(192,57,43,0.12);--red-soft:#fdf2f0;
  --ease-out:cubic-bezier(0.23,1,0.32,1);
  --ease-in-out:cubic-bezier(0.77,0,0.175,1);
  --bg:#eceef4;
  --surface:#ffffff;
  --border:#e2e6f0;
  --text-primary:#181b2e;
  --text-secondary:#626880;
  --text-hint:#9da5be;
  --shadow-card:0 1px 2px rgba(0,0,0,0.04),0 6px 20px rgba(0,0,0,0.06),0 24px 56px rgba(0,0,0,0.05);
}
body{
  font-family:'Sora',system-ui,-apple-system,sans-serif;
  background:var(--bg);min-height:100vh;display:flex;
  align-items:center;justify-content:center;padding:24px;
  -webkit-font-smoothing:antialiased;-moz-osx-font-smoothing:grayscale;
}

/* ── Card ───────────────────────────────────── */
.card{
  background:var(--surface);border-radius:20px;
  box-shadow:var(--shadow-card);
  width:100%;max-width:580px;padding:36px 40px 40px;
  animation:cardFade 380ms var(--ease-out) both;
}
@keyframes cardFade{
  from{opacity:0;transform:translateY(18px)}
  to{opacity:1;transform:translateY(0)}
}

/* ── App header ─────────────────────────────── */
.app-header{display:flex;align-items:center;gap:14px;margin-bottom:8px}
.app-badge{
  width:46px;height:46px;background:var(--red);border-radius:12px;flex-shrink:0;
  display:flex;align-items:center;justify-content:center;
  color:#fff;font-weight:700;font-size:15px;letter-spacing:-0.5px;
  box-shadow:0 4px 12px rgba(192,57,43,0.35),inset 0 1px 0 rgba(255,255,255,0.15);
}
.app-title{font-size:19px;font-weight:700;color:var(--text-primary);letter-spacing:-0.4px;line-height:1.2}
.app-meta{display:flex;align-items:center;gap:7px;margin-top:4px}
.app-org{font-size:12px;color:var(--text-secondary);font-weight:500}
.version-pill{
  display:inline-flex;align-items:center;
  background:#eef2ff;color:#3d5afe;
  font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:20px;letter-spacing:0.2px;
}
.divider{border:none;border-top:1px solid var(--border);margin:22px 0 24px}

/* ── Section ────────────────────────────────── */
.form-section{margin-bottom:22px}
.section-label{display:flex;align-items:center;gap:7px;margin-bottom:14px}
.section-dot{width:6px;height:6px;border-radius:50%;background:var(--red);flex-shrink:0}
.section-name{font-size:10.5px;font-weight:700;color:var(--text-secondary);
              text-transform:uppercase;letter-spacing:0.8px}

/* ── Field ──────────────────────────────────── */
.field{margin-bottom:13px}
.field-label{display:flex;align-items:baseline;gap:5px;
             font-size:12.5px;font-weight:600;color:var(--text-primary);margin-bottom:7px}
.field-hint{font-weight:400;color:var(--text-hint);font-size:11.5px}

/* ── Upload zone ────────────────────────────── */
.upload-zone{
  display:flex;align-items:center;gap:12px;
  width:100%;padding:12px 14px;
  border:1.5px dashed var(--border);border-radius:11px;
  background:#f9fafd;cursor:pointer;position:relative;
  transition:border-color 150ms var(--ease-out),background 150ms var(--ease-out),
             box-shadow 150ms var(--ease-out);
}
@media(hover:hover) and (pointer:fine){
  .upload-zone:hover{border-color:var(--red);background:var(--red-soft);
    box-shadow:0 0 0 3px var(--red-ring)}
}
.upload-zone:focus-within{border-color:var(--red);background:var(--red-soft);
  box-shadow:0 0 0 3px var(--red-ring)}
.upload-zone input[type=file]{
  position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;font-size:0}
.uz-icon{
  width:34px;height:34px;border-radius:9px;background:#edf0f8;flex-shrink:0;
  display:flex;align-items:center;justify-content:center;
  transition:background 150ms var(--ease-out);
}
.upload-zone:focus-within .uz-icon,.upload-zone:hover .uz-icon{background:rgba(192,57,43,0.1)}
.uz-icon svg{width:16px;height:16px;color:var(--text-secondary);transition:color 150ms ease}
.upload-zone:focus-within .uz-icon svg,.upload-zone:hover .uz-icon svg{color:var(--red)}
.uz-text{flex:1;min-width:0}
.uz-name{font-size:13px;font-weight:500;color:var(--text-primary);
         white-space:nowrap;overflow:hidden;text-overflow:ellipsis;display:block}
.uz-name.ph{color:var(--text-hint);font-weight:400}
.uz-ext{font-size:11px;color:var(--text-hint);margin-top:2px;display:block}

/* ── Text inputs ────────────────────────────── */
.row2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.text-input{
  width:100%;padding:10px 13px;
  border:1.5px solid var(--border);border-radius:10px;
  font-size:13px;color:var(--text-primary);
  font-family:'Sora',system-ui,sans-serif;
  background:#f9fafd;outline:none;
  transition:border-color 150ms var(--ease-out),box-shadow 150ms var(--ease-out),
             background 150ms var(--ease-out);
}
.text-input::placeholder{color:var(--text-hint)}
.text-input:focus{border-color:var(--red);background:#fff;
  box-shadow:0 0 0 3px var(--red-ring)}

/* ── Toggle option row ──────────────────────── */
.toggle-row{
  display:flex;align-items:center;gap:10px;
  background:#f7f8fb;border:1.5px solid var(--border);
  border-radius:11px;padding:13px 15px;margin-bottom:14px;cursor:pointer;
  transition:background 150ms var(--ease-out),border-color 150ms var(--ease-out);
}
@media(hover:hover) and (pointer:fine){
  .toggle-row:hover{background:#eef1f9;border-color:#c8d0e5}
}
.toggle-row input[type=checkbox]{width:17px;height:17px;cursor:pointer;
  accent-color:var(--red);flex-shrink:0}
.toggle-label{font-size:13.5px;font-weight:500;color:var(--text-primary);cursor:pointer}

/* ── Primary button ─────────────────────────── */
.btn-primary{
  width:100%;padding:14px;color:#fff;border:none;border-radius:11px;
  font-size:14px;font-weight:600;cursor:pointer;
  font-family:'Sora',system-ui,sans-serif;letter-spacing:0.1px;
  background:linear-gradient(160deg,#cc3d30 0%,#b02e22 100%);
  box-shadow:0 1px 3px rgba(0,0,0,0.14),0 4px 14px rgba(192,57,43,0.32),
             inset 0 1px 0 rgba(255,255,255,0.12);
  transition:transform 160ms var(--ease-out),box-shadow 200ms var(--ease-out),
             background 200ms var(--ease-out);
}
@media(hover:hover) and (pointer:fine){
  .btn-primary:not(:disabled):hover{
    background:linear-gradient(160deg,#c03828 0%,#9e2a1f 100%);
    box-shadow:0 1px 3px rgba(0,0,0,0.14),0 8px 22px rgba(192,57,43,0.42),
               inset 0 1px 0 rgba(255,255,255,0.12);
  }
}
.btn-primary:not(:disabled):active{
  transform:scale(0.97);
  box-shadow:0 1px 2px rgba(0,0,0,0.14),0 2px 6px rgba(192,57,43,0.22);
}
.btn-primary:disabled{background:#cdd2e0;box-shadow:none;cursor:not-allowed}

/* ── Status box ─────────────────────────────── */
#status{
  margin-top:20px;padding:16px 18px;border-radius:13px;
  font-size:13px;display:none;line-height:1.7;
  opacity:0;transform:translateY(6px);
  transition:opacity 220ms var(--ease-out),transform 220ms var(--ease-out);
}
#status.vis{opacity:1;transform:translateY(0)}
.sl{background:#eef2ff;color:#2c5fd4;border:1.5px solid #cdd8ff}
.ss{background:#edfaf3;color:#186840;border:1.5px solid #aae0c4}
.se{background:#fef2f1;color:#9c2418;border:1.5px solid #f5c0bb}

/* ── Download button ────────────────────────── */
.dl-btn{
  display:inline-flex;align-items:center;gap:7px;margin-top:13px;
  padding:10px 18px;background:#186840;color:#fff;border-radius:9px;
  text-decoration:none;font-size:13px;font-weight:600;
  font-family:'Sora',system-ui,sans-serif;
  transition:background 200ms var(--ease-out),transform 160ms var(--ease-out);
}
@media(hover:hover) and (pointer:fine){.dl-btn:hover{background:#145535}}
.dl-btn:active{transform:scale(0.97)}

/* ── Stat cards ─────────────────────────────── */
.stats{display:flex;gap:10px;margin-top:13px;flex-wrap:wrap}
.stat-card{
  background:var(--surface);border:1.5px solid var(--border);border-radius:11px;
  padding:11px 16px;font-size:12px;color:var(--text-secondary);
  flex:1;min-width:110px;text-align:center;
  opacity:0;transform:translateY(8px);
  animation:statFade 280ms var(--ease-out) forwards;
}
.stat-card:nth-child(1){animation-delay:0ms}
.stat-card:nth-child(2){animation-delay:60ms}
.stat-card:nth-child(3){animation-delay:120ms}
@keyframes statFade{to{opacity:1;transform:translateY(0)}}
.stat-num{
  display:block;font-size:24px;font-weight:700;color:var(--red);
  letter-spacing:-0.8px;line-height:1.1;margin-bottom:3px;
}
.stat-num.green{color:#186840}
.stat-num.blue{color:#2c5fd4}

/* ── Warn box ───────────────────────────────── */
.warn-box{
  background:#fffbec;border:1.5px solid #f5e09c;border-radius:8px;
  padding:8px 12px;margin-top:6px;font-size:12px;color:#7a5c0a;
}

/* ── Spinner ────────────────────────────────── */
.spinner{
  display:inline-block;width:13px;height:13px;border:2px solid #c5d5ff;
  border-top-color:#2c5fd4;border-radius:50%;
  animation:spin .7s linear infinite;margin-right:7px;vertical-align:middle;
}
@keyframes spin{to{transform:rotate(360deg)}}

/* ── Picker ─────────────────────────────────── */
.picker-box{
  max-height:200px;overflow-y:auto;border:1.5px solid var(--border);
  border-radius:10px;padding:6px 10px;background:#f9fafd;
}
.prow{
  display:flex;align-items:center;gap:8px;padding:7px 4px;
  font-size:13px;color:var(--text-primary);cursor:pointer;
  border-bottom:1px solid #edf0f8;border-radius:6px;
  transition:background 100ms ease;
}
@media(hover:hover) and (pointer:fine){.prow:hover{background:#eef2ff}}
.prow:last-child{border-bottom:none}
.prow input{width:15px;height:15px;accent-color:var(--red);flex-shrink:0;cursor:pointer}

.phdr{display:flex;align-items:center;justify-content:space-between;margin-bottom:9px}
.toggle-btn{
  font-size:11px;padding:3px 10px;border:1.5px solid var(--border);
  background:var(--surface);color:var(--text-secondary);border-radius:6px;cursor:pointer;
  font-family:'Sora',system-ui,sans-serif;
  transition:background 150ms var(--ease-out),border-color 150ms var(--ease-out),
             color 150ms var(--ease-out),transform 120ms var(--ease-out);
}
@media(hover:hover) and (pointer:fine){
  .toggle-btn:hover{background:var(--red);border-color:var(--red);color:#fff}
}
.toggle-btn:active{transform:scale(0.95)}
.picker-msg{font-size:12.5px;color:var(--text-hint);padding:6px 2px}

.changed-badge{
  display:inline-flex;align-items:center;
  background:#eef2ff;color:#3d5afe;
  font-size:10.5px;font-weight:700;padding:2px 9px;border-radius:20px;margin-left:6px;
}

/* ── Reduced motion ─────────────────────────── */
@media(prefers-reduced-motion:reduce){
  .card{animation:none}
  .stat-card{animation:none;opacity:1;transform:none}
  #status{transition:none}
  .btn-primary,.dl-btn,.toggle-btn{transition:background 150ms ease}
  .btn-primary:not(:disabled):active,.dl-btn:active,.toggle-btn:active{transform:none}
  .upload-zone,.toggle-row,.text-input{transition:none}
}
</style>
</head>
<body>
<div class="card">

  <!-- Header -->
  <div class="app-header">
    <div class="app-badge" aria-hidden="true">HC</div>
    <div>
      <h1 class="app-title">LOA Generator</h1>
      <div class="app-meta">
        <span class="app-org">HC Global Fund Services</span>
        <span class="version-pill">v2 · Change Detection</span>
      </div>
    </div>
  </div>

  <hr class="divider">

  <form id="f" novalidate>

    <!-- Data Files -->
    <div class="form-section">
      <div class="section-label">
        <span class="section-dot" aria-hidden="true"></span>
        <span class="section-name">Data Files</span>
      </div>

      <div class="field">
        <div class="field-label" id="lbl-xl">
          Current Excel File <span class="field-hint">latest investor banking data</span>
        </div>
        <label class="upload-zone" for="xl" aria-labelledby="lbl-xl">
          <div class="uz-icon" aria-hidden="true">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
          </div>
          <div class="uz-text">
            <span class="uz-name ph" id="xl-name">Choose or drop an Excel file</span>
            <span class="uz-ext">.xlsx · .xls</span>
          </div>
          <input type="file" id="xl" accept=".xlsx,.xls" required aria-required="true">
        </label>
      </div>

      <div class="field">
        <div class="field-label" id="lbl-hcg">
          HCGlobal Report <span class="field-hint">previous record — CSV or XLSX</span>
        </div>
        <label class="upload-zone" for="hcg" aria-labelledby="lbl-hcg">
          <div class="uz-icon" aria-hidden="true">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>
          </div>
          <div class="uz-text">
            <span class="uz-name ph" id="hcg-name">Choose or drop a report file</span>
            <span class="uz-ext">.xlsx · .xls · .csv</span>
          </div>
          <input type="file" id="hcg" accept=".xlsx,.xls,.csv" required aria-required="true">
        </label>
      </div>
    </div>

    <hr class="divider">

    <!-- Document Details -->
    <div class="form-section">
      <div class="section-label">
        <span class="section-dot" aria-hidden="true"></span>
        <span class="section-name">Document Details</span>
      </div>
      <div class="row2">
        <div class="field">
          <label class="field-label" for="fn">
            Fund Name <span class="field-hint">after "Re:"</span>
          </label>
          <input class="text-input" type="text" id="fn"
            placeholder="HC Global Opportunity Fund I, LP" required aria-required="true">
        </div>
        <div class="field">
          <label class="field-label" for="dt">
            Date <span class="field-hint">MM/DD/YYYY</span>
          </label>
          <input class="text-input" type="text" id="dt"
            placeholder="05/05/2026" maxlength="10" required aria-required="true"
            inputmode="numeric" autocomplete="off">
        </div>
      </div>
    </div>

    <!-- Changed Accounts Picker (auto-loads when both files selected) -->
    <div id="changed-section" style="display:none">
      <hr class="divider">
      <div class="form-section">
        <div class="section-label">
          <span class="section-dot" aria-hidden="true"></span>
          <span class="section-name">Changed Accounts</span>
          <span id="changed-badge" class="changed-badge" style="display:none"></span>
        </div>
        <div id="changed-loading" class="picker-msg">
          <span class="spinner" aria-hidden="true"></span>Comparing files&hellip;
        </div>
        <div id="changed-body" style="display:none">
          <p style="font-size:12px;color:var(--text-hint);margin-bottom:10px">
            Select which changed accounts to generate LOAs for (all selected by default):
          </p>
          <div class="phdr">
            <span></span>
            <button type="button" id="sel-changed-btn" class="toggle-btn">Deselect All</button>
          </div>
          <div id="changed-list" class="picker-box" role="group" aria-label="Changed accounts"></div>
          <p id="changed-empty" class="picker-msg" style="display:none">No changed accounts detected.</p>
        </div>
      </div>
    </div>

    <!-- Options -->
    <hr class="divider">
    <div class="form-section">
      <div class="section-label">
        <span class="section-dot" aria-hidden="true"></span>
        <span class="section-name">Options</span>
      </div>
      <label class="toggle-row">
        <input type="checkbox" id="incl_unchanged">
        <span class="toggle-label">Also generate LOAs for accounts with <strong>no changes</strong></span>
      </label>
      <div id="unchanged-picker" style="display:none;margin-bottom:20px">
        <div id="picker-loading" class="picker-msg">
          <span class="spinner" aria-hidden="true"></span>Loading unchanged accounts&hellip;
        </div>
        <div id="picker-body" style="display:none">
          <div class="phdr">
            <span style="font-size:12.5px;font-weight:600;color:var(--text-primary)">Select accounts to include:</span>
            <button type="button" id="sel-all-btn" class="toggle-btn">Select All</button>
          </div>
          <div id="acct-list" class="picker-box" role="group" aria-label="Unchanged accounts"></div>
          <p id="picker-empty" class="picker-msg" style="display:none">No unchanged accounts found.</p>
        </div>
      </div>
    </div>

    <button type="submit" class="btn-primary" id="gb">
      &#9889; Detect Changes &amp; Generate LOAs
    </button>

  </form>

  <div id="status" role="status" aria-live="polite"></div>
</div>

<script>
const form=document.getElementById('f'),st=document.getElementById('status'),gb=document.getElementById('gb');
const inclCb=document.getElementById('incl_unchanged');
const changedSection=document.getElementById('changed-section');
const changedLoading=document.getElementById('changed-loading');
const changedBody=document.getElementById('changed-body');
const changedList=document.getElementById('changed-list');
const changedEmpty=document.getElementById('changed-empty');
const changedBadge=document.getElementById('changed-badge');
const selChangedBtn=document.getElementById('sel-changed-btn');
const picker=document.getElementById('unchanged-picker');
const pickerLoading=document.getElementById('picker-loading');
const pickerBody=document.getElementById('picker-body');
const acctList=document.getElementById('acct-list');
const pickerEmpty=document.getElementById('picker-empty');
const selAllBtn=document.getElementById('sel-all-btn');

let cachedChanged=[], cachedUnchanged=[];

// Date auto-format
document.getElementById('dt').addEventListener('input',function(e){
  let v=e.target.value.replace(/[^0-9]/g,'');
  if(v.length>=3)v=v.slice(0,2)+'/'+v.slice(2);
  if(v.length>=6)v=v.slice(0,5)+'/'+v.slice(5,9);
  e.target.value=v;
});

// File name display helpers
function wireFileInput(inputId, nameId, defaultText){
  const inp=document.getElementById(inputId);
  const disp=document.getElementById(nameId);
  inp.addEventListener('change',function(){
    if(inp.files[0]){
      disp.textContent=inp.files[0].name;
      disp.classList.remove('ph');
    } else {
      disp.textContent=defaultText;
      disp.classList.add('ph');
    }
  });
}
wireFileInput('xl','xl-name','Choose or drop an Excel file');
wireFileInput('hcg','hcg-name','Choose or drop a report file');

function show(cls,html){
  st.className=cls;st.innerHTML=html;
  st.style.display='block';
  st.offsetHeight;
  st.classList.add('vis');
}

// Picker helpers
function renderList(names,listEl,emptyEl,preChecked){
  listEl.innerHTML='';
  if(!names.length){emptyEl.style.display='block';return;}
  emptyEl.style.display='none';
  names.forEach(function(name){
    const lbl=document.createElement('label');
    lbl.className='prow';
    const cb=document.createElement('input');
    cb.type='checkbox';cb.value=name;cb.checked=preChecked;
    lbl.appendChild(cb);
    const sp=document.createElement('span');sp.textContent=name;
    lbl.appendChild(sp);
    listEl.appendChild(lbl);
  });
}
function allChecked(listEl){
  const cbs=listEl.querySelectorAll('input[type=checkbox]');
  return cbs.length>0&&[...cbs].every(c=>c.checked);
}
function syncToggleBtn(listEl,btn){
  const cbs=listEl.querySelectorAll('input[type=checkbox]');
  btn.textContent=(!cbs.length||allChecked(listEl))?'Deselect All':'Select All';
}
function doToggle(listEl,btn){
  const yes=allChecked(listEl);
  listEl.querySelectorAll('input[type=checkbox]').forEach(c=>{c.checked=!yes;});
  btn.textContent=yes?'Select All':'Deselect All';
}

selChangedBtn.addEventListener('click',function(){doToggle(changedList,selChangedBtn);});
changedList.addEventListener('change',function(){syncToggleBtn(changedList,selChangedBtn);});
selAllBtn.addEventListener('click',function(){doToggle(acctList,selAllBtn);});
acctList.addEventListener('change',function(){syncToggleBtn(acctList,selAllBtn);});

async function loadPreviews(){
  const xl=document.getElementById('xl').files[0];
  const hcg=document.getElementById('hcg').files[0];
  if(!xl||!hcg) return;

  changedSection.style.display='block';
  changedLoading.style.display='block';
  changedBody.style.display='none';
  changedBadge.style.display='none';
  changedList.innerHTML='';
  changedEmpty.style.display='none';

  if(inclCb.checked){
    picker.style.display='block';
    pickerLoading.style.display='block';
    pickerBody.style.display='none';
    acctList.innerHTML='';
  }

  const fd=new FormData();
  fd.append('excel',xl);fd.append('hcg',hcg);fd.append('hcg_filename',hcg.name);
  try{
    const res=await fetch('/preview_accounts',{method:'POST',body:fd});
    const data=await res.json();
    if(data.error){
      changedLoading.style.display='none';
      changedEmpty.textContent='Error: '+data.error;
      changedEmpty.style.display='block';
      changedBody.style.display='block';
      return;
    }
    cachedChanged=data.changed||[];
    cachedUnchanged=data.unchanged||[];
    changedLoading.style.display='none';
    renderList(cachedChanged,changedList,changedEmpty,true);
    if(cachedChanged.length){
      changedBadge.textContent=cachedChanged.length+' account'+(cachedChanged.length>1?'s':'')+' with changes';
      changedBadge.style.display='inline-flex';
    }
    syncToggleBtn(changedList,selChangedBtn);
    changedBody.style.display='block';
    if(inclCb.checked){
      pickerLoading.style.display='none';
      renderList(cachedUnchanged,acctList,pickerEmpty,true);
      syncToggleBtn(acctList,selAllBtn);
      pickerBody.style.display='block';
    }
  }catch(err){
    changedLoading.style.display='none';
    changedEmpty.textContent='Could not load: '+err.message;
    changedEmpty.style.display='block';
    changedBody.style.display='block';
    if(inclCb.checked){pickerLoading.style.display='none';pickerBody.style.display='block';}
  }
}

document.getElementById('xl').addEventListener('change',loadPreviews);
document.getElementById('hcg').addEventListener('change',loadPreviews);

inclCb.addEventListener('change',function(){
  if(inclCb.checked){
    picker.style.display='block';
    if(cachedUnchanged.length>0||cachedChanged.length>0){
      pickerLoading.style.display='none';
      renderList(cachedUnchanged,acctList,pickerEmpty,true);
      syncToggleBtn(acctList,selAllBtn);
      pickerBody.style.display='block';
    } else {
      loadPreviews();
    }
  } else {
    picker.style.display='none';
    acctList.innerHTML='';
  }
});

form.addEventListener('submit',async function(e){
  e.preventDefault();
  const xl=document.getElementById('xl').files[0];
  const hcg=document.getElementById('hcg').files[0];
  const fn=document.getElementById('fn').value.trim();
  const dt=document.getElementById('dt').value.trim();
  const incl=inclCb.checked;
  if(!xl||!hcg||!fn||!dt){show('se','Please fill in all fields and select both files.');return;}
  if(!/^\d{2}\/\d{2}\/\d{4}$/.test(dt)){show('se','Date must be MM/DD/YYYY format.');return;}

  let selectedChanged=[];
  changedList.querySelectorAll('input[type=checkbox]:checked').forEach(function(cb){
    selectedChanged.push(cb.value);
  });
  let selectedUnchanged=[];
  if(incl){
    acctList.querySelectorAll('input[type=checkbox]:checked').forEach(function(cb){
      selectedUnchanged.push(cb.value);
    });
  }

  gb.disabled=true;
  show('sl','<span class="spinner" aria-hidden="true"></span>Comparing records and generating LOAs&hellip;');
  const fd=new FormData();
  fd.append('excel',xl);fd.append('hcg',hcg);fd.append('hcg_filename',hcg.name);
  fd.append('fund_name',fn);fd.append('date',dt);
  fd.append('include_unchanged',incl?'1':'0');
  fd.append('selected_changed',JSON.stringify(selectedChanged));
  fd.append('selected_unchanged',JSON.stringify(selectedUnchanged));
  try{
    const res=await fetch('/generate',{method:'POST',body:fd});
    const data=await res.json();
    if(data.error){show('se','&#9888; '+data.error);return;}
    let html='<div class="stats">';
    html+='<div class="stat-card"><span class="stat-num">'+data.changed_generated+'</span>Changed LOAs</div>';
    html+='<div class="stat-card"><span class="stat-num green">'+data.unchanged+'</span>Unchanged</div>';
    if(data.new_accts>0)html+='<div class="stat-card"><span class="stat-num blue">'+data.new_accts+'</span>New</div>';
    html+='</div>';
    html+='<a class="dl-btn" href="/download/'+data.token+'" download="LOA_Package.zip">';
    html+='<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>';
    html+='Download LOA_Package.zip</a>';
    html+='<div style="font-size:11.5px;color:var(--text-hint);margin-top:7px">Includes PDFs + <strong>Change_Report.xlsx</strong></div>';
    if(data.warnings&&data.warnings.length){
      html+='<div style="margin-top:12px"><strong style="font-size:12px">&#9888; Warnings:</strong>';
      data.warnings.forEach(w=>{html+='<div class="warn-box">'+w+'</div>';});
      html+='</div>';
    }
    show('ss',html);
  }catch(err){show('se','Error: '+err.message);}
  finally{gb.disabled=false;}
});
</script>
</body>
</html>"""

# ── Store ─────────────────────────────────────────────────────────────────────
_zip_store = {}

# ── Helpers ───────────────────────────────────────────────────────────────────
def clean(val):
    if val is None: return ""
    try:
        if isinstance(val, float) and math.isnan(val): return ""
    except Exception: pass
    s = str(val).strip()
    return "" if s.lower() in ("nan","none","-","n/a") else s

def safe_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "_", name)

def normalize_cmp(val):
    v = clean(val)
    if re.match(r'^\d+\.0$', v):
        v = v[:-2]
    return v.lower()

def get_display_name(excel_row):
    """Unique display name for an Excel row.
    For custodian accounts (e.g. AltoIRA Custodian) where Contact Name
    disambiguates individual investors, returns 'Account Name — Contact Name'."""
    acct    = clean(excel_row.get("Account Name", ""))
    contact = clean(excel_row.get("Contact Name", ""))
    return f"{acct} — {contact}" if contact else acct

def get_match_keys(excel_row):
    """Priority-ordered list of HCG lookup keys for this Excel row.
    Tries composite and contact-name keys first so custodian accounts
    (multiple rows sharing the same Account Name) match correctly."""
    acct    = clean(excel_row.get("Account Name", ""))
    contact = clean(excel_row.get("Contact Name", ""))
    keys = []
    if acct and contact:
        # e.g. "altoira custodian — jane doe" or "altoira custodian jane doe"
        keys.append(normalize_cmp(f"{acct} {contact}"))
        keys.append(normalize_cmp(contact))      # HCG may store investor name only
    keys.append(normalize_cmp(acct))
    return [k for k in keys if k]

def build_hcg_lookup(df_hcg):
    """Build a dict {normalized_key: row_dict} from an HCG DataFrame.
    Each row is indexed under both its account name and investor name so
    that custodian accounts can be found via either identifier."""
    lookup = {}
    for _, row in df_hcg.iterrows():
        rdict = dict(row)
        for key in [
            normalize_cmp(rdict.get("hcg_acct_name", "")),
            normalize_cmp(rdict.get("hcg_investor_name", "")),
        ]:
            if key and key not in lookup:
                lookup[key] = rdict
    return lookup

def parse_multipart(headers, rfile):
    content_type = headers.get("Content-Type","")
    m = re.search(r"boundary=([^\s;]+)", content_type)
    if not m: return {}
    boundary = m.group(1).strip('"').strip("'")
    content_length = int(headers.get("Content-Length", 0))
    body = rfile.read(content_length)
    delimiter = ("--" + boundary).encode()
    parts = {}
    for chunk in body.split(delimiter)[1:]:
        if chunk.strip() in (b"", b"--", b"--\r\n"): continue
        if chunk.startswith(b"\r\n"): chunk = chunk[2:]
        sep = chunk.find(b"\r\n\r\n")
        if sep == -1: continue
        hdr = chunk[:sep].decode("utf-8", errors="replace")
        body_part = chunk[sep+4:]
        if body_part.endswith(b"\r\n"): body_part = body_part[:-2]
        nm = re.search(r'name="([^"]+)"', hdr)
        fm = re.search(r'filename="([^"]*)"', hdr)
        if not nm: continue
        name = nm.group(1)
        parts[name] = body_part if fm else body_part.decode("utf-8", errors="replace")
    return parts

# ── HCG Report Reader – fuzzy column matching ─────────────────────────────────
def _norm_col(s):
    """Normalise a column header for fuzzy matching: lowercase, collapse whitespace/punctuation."""
    s = re.sub(r'\s+', ' ', str(s)).strip().lower()
    s = re.sub(r'[/\-#().\']', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

# (target_internal_name, [list of known pattern variants — most specific first])
# Patterns are matched against the normalised actual column names.
HCG_COL_PATTERNS = [
    # Account identity
    ("hcg_acct_name",    ["bank account name", "bank acct name", "beneficiary name",
                          "account name", "acct name"]),
    ("hcg_investor_name",["investor name", "client name", "investor"]),
    # Account details
    ("hcg_acct_number",  ["account number", "acct number", "account no", "acct no",
                          "account num", "acct num", "account nbr"]),
    ("hcg_aba",          ["aba routing", "aba  routing", "routing number", "routing no",
                          "aba number", "routing", "aba"]),
    ("hcg_swift",        ["swift code", "swift bic", "bic code", "swift  bic",
                          "swift", "bic"]),
    ("hcg_iban",         ["iban code", "iban number", "iban"]),
    ("hcg_ffc",          ["ffc instruction", "for further credit", "further credit",
                          "ffc to", "ffc"]),
    # Type / payment columns — note: "hcg_acct_type" listed before "hcg_payment_type"
    # so a column literally named "Type" resolves to account type first.
    ("hcg_acct_type",    ["account type", "acct type", "type"]),
    ("hcg_payment_type", ["payment type ach wire", "payment type ach  wire",
                          "payment type", "payment method", "ach wire", "ach  wire"]),
]

def _col_match_score(norm_col, norm_pat):
    """
    Return a match score (>0 = match):
      exact        → len(pattern) × 10   (highest)
      pat in col   → len(pattern)         (descriptive column: "ABA Routing Number")
      col in pat   → len(col)             (abbreviated column: "ABA" inside "aba routing")
    Score 0 = no match.
    """
    if not norm_pat:
        return 0
    if norm_col == norm_pat:
        return len(norm_pat) * 10
    if norm_pat in norm_col:
        return len(norm_pat)
    if len(norm_col) >= 3 and norm_col in norm_pat:
        return len(norm_col)
    return 0

def _map_hcg_columns(actual_cols):
    """
    Given a list of actual (raw) column names, return a rename dict
    {actual_col: internal_name} using fuzzy scoring.
    Each internal name is assigned at most once (greedy, highest score wins;
    ties broken by target list order).
    """
    norm_cols = [(c, _norm_col(c)) for c in actual_cols]
    assigned_targets = set()
    rename = {}

    # Score every (actual_col, target) pair; keep best per actual_col
    # Structure: {actual_col: (best_score, target_index, target_name)}
    best = {}
    for actual_col, norm_col in norm_cols:
        for t_idx, (target, patterns) in enumerate(HCG_COL_PATTERNS):
            top_score = max(_col_match_score(norm_col, _norm_col(p)) for p in patterns)
            if top_score > 0:
                prev = best.get(actual_col, (0, 999, None))
                # Prefer higher score; tie-break by earlier target in list (lower t_idx)
                if (top_score, -t_idx) > (prev[0], -prev[1]):
                    best[actual_col] = (top_score, t_idx, target)

    # Sort by score descending so best-matched columns are assigned first
    ranked = sorted(best.items(), key=lambda x: (-x[1][0], x[1][1]))
    for actual_col, (score, t_idx, target) in ranked:
        if target not in assigned_targets:
            rename[actual_col] = target
            assigned_targets.add(target)

    return rename

def read_hcg_report(file_bytes, filename):
    """Read HCG CSV or XLSX, return normalised DataFrame with standard internal col names."""
    if filename.lower().endswith(".csv"):
        df = None
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), dtype=str,
                                 encoding=enc, sep=None, engine="python")
                break
            except Exception:
                continue
        if df is None:
            raise ValueError("Could not read CSV with any known encoding/separator.")
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)

    # Collapse embedded newlines / extra spaces in headers
    df.columns = [re.sub(r'\s+', ' ', str(c)).strip() for c in df.columns]

    # Fuzzy-map actual headers → internal names
    rename = _map_hcg_columns(list(df.columns))
    df = df.rename(columns=rename)

    # Guarantee every internal column exists (fill missing with empty string)
    for col in ["hcg_acct_name", "hcg_investor_name", "hcg_acct_number",
                "hcg_aba", "hcg_swift", "hcg_iban", "hcg_ffc", "hcg_acct_type"]:
        if col not in df.columns:
            df[col] = ""

    # Fall back to investor name if bank account name was not mapped / is blank
    df["hcg_acct_name"] = df.apply(
        lambda r: clean(r["hcg_acct_name"]) or clean(r["hcg_investor_name"]), axis=1)
    return df

# ── Field comparison ──────────────────────────────────────────────────────────
# Returns list of (field_label, old_value, new_value) for any differences
COMPARE_FIELDS = [
    # (label, excel_col, hcg_col)   – ABA handled separately
    ("Account Number",               "Account Number",  "hcg_acct_number"),
    ("Account Type",                 "Account Type",    "hcg_acct_type"),
    ("FFC Instruction (if applicable)", "FFC Instruction", "hcg_ffc"),
    ("Swift code (if applicable)",   "Swift Code",      "hcg_swift"),
    ("Iban code (if applicable)",    "IBAN Code",       "hcg_iban"),
]

def get_excel_aba_for_type(excel_row, payment_type=None):
    """Return the Excel routing number that matches the given payment type.
    payment_type: string from HCG record (e.g. 'Wire', 'ACH', 'ACH/Wire', blank).
    Falls back to Wire→ACH when type is unknown."""
    wire = clean(excel_row.get("Routing Number (Wire)", ""))
    ach  = clean(excel_row.get("Routing Number (ACH)",  ""))
    pt   = (payment_type or "").lower()
    if "wire" in pt:
        return wire if wire else ach   # Wire preferred; ACH as last resort
    if "ach" in pt:
        return ach  if ach  else wire  # ACH preferred; Wire as last resort
    return wire if wire else ach       # Unknown type: prefer Wire

def compare_records(excel_row, hcg_row):
    """Return (diffs, aba_warning).
    diffs       : list of (label, old_val, new_val)  –  empty means no changes
    aba_warning : str or None  –  set when the Excel routing type doesn't
                  cleanly match the HCG payment type"""
    diffs = []
    aba_warning = None

    # Determine HCG payment type (e.g. "Wire", "ACH", "ACH/Wire")
    hcg_pt  = clean(hcg_row.get("hcg_payment_type", ""))
    old_aba = clean(hcg_row.get("hcg_aba", ""))
    pt_low  = hcg_pt.lower()

    wire = clean(excel_row.get("Routing Number (Wire)", ""))
    ach  = clean(excel_row.get("Routing Number (ACH)",  ""))

    # Pick the Excel routing column that corresponds to the HCG payment type
    new_aba = get_excel_aba_for_type(excel_row, hcg_pt)

    # Flag when the preferred column is missing and we had to fall back
    if "wire" in pt_low and not wire and ach:
        aba_warning = (f"HCG payment type is <b>Wire</b> but Excel has no Wire routing "
                       f"number — ACH number ({ach}) was used for comparison.")
    elif "ach" in pt_low and not ach and wire:
        aba_warning = (f"HCG payment type is <b>ACH</b> but Excel has no ACH routing "
                       f"number — Wire number ({wire}) was used for comparison.")

    if normalize_cmp(new_aba) != normalize_cmp(old_aba):
        diffs.append(("ABA Routing Number", old_aba, new_aba))

    # Other fields
    for label, ecol, hcol in COMPARE_FIELDS:
        nv = clean(excel_row.get(ecol, ""))
        ov = clean(hcg_row.get(hcol, ""))
        if normalize_cmp(nv) != normalize_cmp(ov):
            diffs.append((label, ov, nv))

    return diffs, aba_warning

# ── PDF Generation ────────────────────────────────────────────────────────────
def generate_loa_pdf(excel_row, hcg_row, fund_name, date_str):
    """
    excel_row : dict of current (new) data
    hcg_row   : dict of old HCG data  (None = new/unchanged account)
    """
    buf = io.BytesIO()
    W, H = letter
    c = rl_canvas.Canvas(buf, pagesize=letter)
    LEFT, RIGHT, Y = 72, W - 72, H - 50

    # Logo
    logo_bytes = base64.b64decode(LOGO_B64)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(logo_bytes); tmp_path = tmp.name
    try:
        c.drawImage(tmp_path, LEFT, Y - 28, width=100, height=28,
                    preserveAspectRatio=True, mask="auto")
    finally:
        os.unlink(tmp_path)
    Y -= 52

    # Address
    c.setFont("Helvetica", 9.5)
    for line in ["HC Global Fund Services, LLC","312 Sutter Street","Suite 601","San Francisco, CA 94108"]:
        c.drawString(LEFT, Y, line); Y -= 13
    Y -= 8

    # Re:
    c.setFont("Helvetica-Bold", 10)
    c.drawString(LEFT, Y, "Re: " + fund_name); Y -= 20

    # Paragraph
    c.setFont("Helvetica", 10)
    c.drawString(LEFT, Y, "Please update the following information within the EquityZen account related to")
    Y -= 16

    # (circle one) line 1
    x0 = LEFT + 62
    c.drawString(LEFT, Y, "(")
    c.setFont("Helvetica-Oblique", 10); c.drawString(LEFT+7, Y, "circle one"); c.setFont("Helvetica", 10)
    c.drawString(x0, Y, ") Remove / ")
    x1 = x0 + c.stringWidth(") Remove / ","Helvetica",10)
    c.setFont("Helvetica-Bold",10); c.drawString(x1, Y, "Change")
    tw = c.stringWidth("Change","Helvetica-Bold",10); c.line(x1,Y-1,x1+tw,Y-1)
    c.setFont("Helvetica",10); c.drawString(x1+tw, Y, " / Add"); Y -= 14

    # (circle one) line 2
    c.drawString(LEFT, Y, "(")
    c.setFont("Helvetica-Oblique",10); c.drawString(LEFT+7,Y,"circle one"); c.setFont("Helvetica",10)
    part = ") Mailing Address / Email Address / Phone # / Authorized Recipient / "
    c.drawString(x0, Y, part)
    x2 = x0 + c.stringWidth(part,"Helvetica",10)
    c.setFont("Helvetica-Bold",10); c.drawString(x2, Y, "Wire Instructions")
    tw2 = c.stringWidth("Wire Instructions","Helvetica-Bold",10); c.line(x2,Y-1,x2+tw2,Y-1)
    c.setFont("Helvetica",10); Y -= 14
    c.drawString(LEFT, Y, "/ Investor Name  If Other, please indicate: _____________________"); Y -= 18

    # Helper: draw a labelled field with text wrapping
    def draw_field(canvas, indent, y, label, value, right_edge):
        canvas.setFont("Helvetica-Bold", 10)
        lw = canvas.stringWidth(label,"Helvetica-Bold",10)
        canvas.drawString(indent, y, label)
        if not value:
            return y - 13
        canvas.setFont("Helvetica", 10)
        if canvas.stringWidth(value,"Helvetica",10) <= right_edge - indent - lw - 6:
            canvas.drawString(indent + lw + 4, y, value)
            return y - 13
        y -= 13
        words = value.split()
        buf = ""
        for w in words:
            test = (buf + " " + w).strip()
            if canvas.stringWidth(test,"Helvetica",10) < right_edge - indent - 12:
                buf = test
            else:
                canvas.drawString(indent+12, y, buf); y -= 13; buf = w
        if buf:
            canvas.drawString(indent+12, y, buf); y -= 13
        return y

    INDENT = LEFT + 54

    # ── "to be removed or changed from" ──
    c.setFont("Helvetica", 10)
    c.drawString(LEFT, Y, "to be removed or changed from:"); Y -= 15

    if hcg_row:
        old_acct_name = clean(hcg_row.get("hcg_acct_name",""))
        old_acct_num  = clean(hcg_row.get("hcg_acct_number",""))
        old_aba       = clean(hcg_row.get("hcg_aba",""))
        old_type      = clean(hcg_row.get("hcg_acct_type",""))
        old_ffc       = clean(hcg_row.get("hcg_ffc",""))
        old_swift     = clean(hcg_row.get("hcg_swift",""))
        old_iban      = clean(hcg_row.get("hcg_iban",""))
    else:
        old_acct_name = old_acct_num = old_aba = old_type = old_ffc = old_swift = old_iban = ""

    from_fields = [
        ("Account Name:",                    old_acct_name),
        ("Account Number:",                  old_acct_num),
        ("ABA Routing Number:",              old_aba),
        ("Account Type:",                    old_type),
        ("FFC Instruction (if applicable):", old_ffc),
        ("Swift code (if applicable):",      old_swift),
        ("Iban code (if applicable):",       old_iban),
    ]
    for label, val in from_fields:
        Y = draw_field(c, INDENT, Y, label, val, RIGHT)
    Y -= 8

    # ── "to be added or changed to" ──
    c.setFont("Helvetica",10)
    c.drawString(LEFT, Y, "to be added or changed to:"); Y -= 15

    new_acct_name = clean(excel_row.get("Account Name",""))
    new_acct_num  = clean(excel_row.get("Account Number",""))
    hcg_pt        = clean(hcg_row.get("hcg_payment_type","")) if hcg_row else ""
    new_aba       = get_excel_aba_for_type(excel_row, hcg_pt)
    new_type      = clean(excel_row.get("Account Type",""))
    new_ffc       = clean(excel_row.get("FFC Instruction",""))
    new_swift     = clean(excel_row.get("Swift Code",""))
    new_iban      = clean(excel_row.get("IBAN Code",""))

    to_fields = [
        ("Account Name:",                    new_acct_name),
        ("Account Number:",                  new_acct_num),
        ("ABA Routing Number:",              new_aba),
        ("Account Type:",                    new_type),
        ("FFC Instruction (if applicable):", new_ffc),
        ("Swift code (if applicable):",      new_swift),
        ("Iban code (if applicable):",       new_iban),
    ]
    for label, val in to_fields:
        Y = draw_field(c, INDENT, Y, label, val, RIGHT)
    Y -= 10

    # Approval
    c.setFont("Helvetica",10); c.drawString(LEFT, Y, "Approval of Change:"); Y -= 16
    bx,by,bs = LEFT,Y-2,10
    c.rect(bx,by,bs,bs)
    c.setLineWidth(1.2); c.line(bx+2,by+4,bx+4,by+2); c.line(bx+4,by+2,bx+8,by+7); c.setLineWidth(1)
    c.setFont("Helvetica",10)
    c.drawString(bx+bs+5, Y, "I hereby approve the update of my information as detailed above.")
    Y -= 22

    c.setFont("Helvetica-Bold",10); c.drawString(LEFT, Y, new_acct_name); Y -= 13
    c.setFont("Helvetica",9);       c.drawString(LEFT, Y, "Name");          Y -= 20
    c.setFont("Helvetica-Bold",10); c.drawString(LEFT, Y, date_str);        Y -= 13
    c.setFont("Helvetica",9);       c.drawString(LEFT, Y, "Date")

    c.save(); buf.seek(0)
    return buf.read()

# ── Change Report (XLSX) ──────────────────────────────────────────────────────
def generate_change_report(changed_list, unchanged_list, new_list,
                           generated_changed=None, generated_unchanged=None):
    """
    changed_list       : [(display_name, [(field, old, new), ...]), ...]
    unchanged_list     : [display_name, ...]
    new_list           : [display_name, ...]
    generated_changed  : set of display_names that had a changed LOA written
    generated_unchanged: set of display_names that had an unchanged LOA written
    Returns bytes of an XLSX workbook.
    """
    generated_changed  = generated_changed  or set()
    generated_unchanged= generated_unchanged or set()
    wb = Workbook()

    # ── Sheet 1: Changes Detail ──
    ws1 = wb.active
    ws1.title = "Changes Detail"
    hdr_fill = PatternFill("solid", fgColor="C0392B")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(["Account Name","Field Changed",
                                  "Previous Value (HCG Report)","New Value (Excel)"], 1):
        cell = ws1.cell(row=1, column=col, value=title)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws1.row_dimensions[1].height = 20

    changed_fill = PatternFill("solid", fgColor="FDECEA")
    r = 2
    for acct_name, diffs in changed_list:
        for field, old_val, new_val in diffs:
            ws1.cell(r,1,acct_name).fill = changed_fill
            ws1.cell(r,2,field).fill = changed_fill
            c_old = ws1.cell(r,3,old_val if old_val else "(blank)")
            c_old.fill = changed_fill; c_old.font = Font(color="922B21")
            c_new = ws1.cell(r,4,new_val if new_val else "(blank)")
            c_new.fill = changed_fill; c_new.font = Font(bold=True, color="1E6B36")
            r += 1

    for col, w in zip("ABCD", [28, 30, 45, 45]):
        ws1.column_dimensions[col].width = w

    # ── Sheet 2: All Accounts Summary ──
    ws2 = wb.create_sheet("All Accounts Summary")
    for col, title in enumerate(["Account Name","Status","Fields Changed"], 1):
        cell = ws2.cell(1, col, title)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 20

    changed_names = {a for a, _ in changed_list}
    r = 2
    for acct_name, diffs in changed_list:
        fields_changed = ", ".join(d[0] for d in diffs)
        ws2.cell(r,1,acct_name); ws2.cell(r,2,"Changed"); ws2.cell(r,3,fields_changed)
        for col in range(1,4):
            ws2.cell(r,col).fill = PatternFill("solid", fgColor="FDECEA")
        r += 1
    for acct_name in unchanged_list:
        ws2.cell(r,1,acct_name); ws2.cell(r,2,"Unchanged"); ws2.cell(r,3,"—")
        for col in range(1,4):
            ws2.cell(r,col).fill = PatternFill("solid", fgColor="E9F7EF")
        r += 1
    for acct_name in new_list:
        ws2.cell(r,1,acct_name); ws2.cell(r,2,"New Account (not in HCG report)"); ws2.cell(r,3,"—")
        for col in range(1,4):
            ws2.cell(r,col).fill = PatternFill("solid", fgColor="EBF5FB")
        r += 1
    for col, w in zip("ABC", [28, 32, 50]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: LOAs Generated ──
    ws3 = wb.create_sheet("LOAs Generated")
    for col, title in enumerate(["Account Name", "Type", "Fields Changed"], 1):
        cell = ws3.cell(1, col, title)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 20

    r = 2
    for acct_name, diffs in changed_list:
        if acct_name not in generated_changed:
            continue
        fields_changed = ", ".join(d[0] for d in diffs)
        ws3.cell(r,1,acct_name); ws3.cell(r,2,"Changed"); ws3.cell(r,3,fields_changed)
        for col in range(1,4):
            ws3.cell(r,col).fill = PatternFill("solid", fgColor="FDECEA")
        r += 1
    for acct_name in new_list:
        ws3.cell(r,1,acct_name); ws3.cell(r,2,"New Account"); ws3.cell(r,3,"—")
        for col in range(1,4):
            ws3.cell(r,col).fill = PatternFill("solid", fgColor="EBF5FB")
        r += 1
    for acct_name in unchanged_list:
        if acct_name not in generated_unchanged:
            continue
        ws3.cell(r,1,acct_name); ws3.cell(r,2,"Unchanged"); ws3.cell(r,3,"—")
        for col in range(1,4):
            ws3.cell(r,col).fill = PatternFill("solid", fgColor="E9F7EF")
        r += 1
    if r == 2:   # nothing generated
        ws3.cell(2,1,"(No LOAs generated)")
    for col, w in zip("ABC", [28, 16, 50]):
        ws3.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read()

# ── HTTP Handler ──────────────────────────────────────────────────────────────
class LOAHandler(BaseHTTPRequestHandler):
    def log_message(self, *args): pass

    def do_GET(self):
        if self.path == "/":
            body = HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type","text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers(); self.wfile.write(body)
        elif self.path.startswith("/download/"):
            token = self.path.split("/")[-1]
            if token in _zip_store:
                data = _zip_store[token]
                self.send_response(200)
                self.send_header("Content-Type","application/zip")
                self.send_header("Content-Disposition",'attachment; filename="LOA_Package.zip"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers(); self.wfile.write(data)
            else:
                self._json(404, {"error":"Download not found."})
        else:
            self._json(404, {"error":"Not found."})

    def do_POST(self):
        if self.path == "/preview_accounts":
            self._handle_preview_accounts(); return
        if self.path != "/generate":
            self._json(404, {"error":"Not found."}); return

        form = parse_multipart(self.headers, self.rfile)

        def fstr(k):
            v = form.get(k,"")
            return (v.decode("utf-8","replace") if isinstance(v, bytes) else v).strip()

        fund_name         = fstr("fund_name")
        date_val          = fstr("date")
        hcg_filename      = fstr("hcg_filename")
        include_unchanged = fstr("include_unchanged") == "1"
        try:
            selected_changed   = set(json.loads(fstr("selected_changed")   or "[]"))
        except Exception:
            selected_changed   = set()
        try:
            selected_unchanged = set(json.loads(fstr("selected_unchanged") or "[]"))
        except Exception:
            selected_unchanged = set()

        if not fund_name or not date_val:
            self._json(400,{"error":"fund_name and date are required."}); return
        if not re.match(r"^\d{2}/\d{2}/\d{4}$", date_val):
            self._json(400,{"error":"Date must be MM/DD/YYYY."}); return

        excel_bytes = form.get("excel")
        hcg_bytes   = form.get("hcg")
        if not excel_bytes or not isinstance(excel_bytes, bytes):
            self._json(400,{"error":"No Excel file uploaded."}); return
        if not hcg_bytes or not isinstance(hcg_bytes, bytes):
            self._json(400,{"error":"No HCG report file uploaded."}); return

        # Read Excel
        try:
            df_xl = pd.read_excel(io.BytesIO(excel_bytes), dtype=str)
            df_xl.columns = [c.strip() for c in df_xl.columns]
        except Exception as e:
            self._json(400,{"error":"Could not read Excel: "+str(e)}); return

        # Read HCG report
        try:
            df_hcg = read_hcg_report(hcg_bytes, hcg_filename or "report.csv")
        except Exception as e:
            self._json(400,{"error":"Could not read HCG report: "+str(e)}); return

        # Build HCG lookup (keyed on account name AND investor name for custodian accounts)
        hcg_lookup = build_hcg_lookup(df_hcg)

        warnings           = []
        changed_list       = []   # [(display_name, diffs)]
        unchanged_list     = []   # [display_name]
        new_list           = []   # [display_name]
        generated_changed  = set()
        generated_unchanged= set()
        zip_buf            = io.BytesIO()

        with zipfile.ZipFile(zip_buf,"w",zipfile.ZIP_DEFLATED) as zf:
            for idx, excel_row in df_xl.iterrows():
                excel_dict   = dict(excel_row)
                acct         = clean(excel_row.get("Account Name",""))
                if not acct:
                    continue
                display_name = get_display_name(excel_dict)

                # Multi-key lookup: try composite/contact name first
                hcg_row = None
                for key in get_match_keys(excel_dict):
                    if key in hcg_lookup:
                        hcg_row = hcg_lookup[key]
                        break

                if hcg_row is None:
                    # New account – not in HCG report
                    new_list.append(display_name)
                    try:
                        pdf = generate_loa_pdf(excel_dict, None, fund_name, date_val)
                        zf.writestr("new/" + safe_filename(display_name) + "_LOA.pdf", pdf)
                    except Exception as e:
                        warnings.append("Row "+str(idx+2)+" <b>"+display_name+"</b>: PDF error: "+str(e))
                    continue

                diffs, aba_warn = compare_records(excel_dict, hcg_row)
                if aba_warn:
                    warnings.append("<b>"+display_name+"</b>: "+aba_warn)

                if diffs:
                    # Changed account
                    changed_list.append((display_name, diffs))
                    # Generate LOA if no explicit selection OR this account is selected
                    if not selected_changed or display_name in selected_changed:
                        try:
                            pdf = generate_loa_pdf(excel_dict, hcg_row, fund_name, date_val)
                            zf.writestr("changed/" + safe_filename(display_name) + "_LOA.pdf", pdf)
                            generated_changed.add(display_name)
                        except Exception as e:
                            warnings.append("Row "+str(idx+2)+" <b>"+display_name+"</b>: PDF error: "+str(e))
                else:
                    # Unchanged account
                    unchanged_list.append(display_name)
                    if include_unchanged and (not selected_unchanged or display_name in selected_unchanged):
                        try:
                            pdf = generate_loa_pdf(excel_dict, hcg_row, fund_name, date_val)
                            zf.writestr("unchanged/" + safe_filename(display_name) + "_LOA.pdf", pdf)
                            generated_unchanged.add(display_name)
                        except Exception as e:
                            warnings.append("Row "+str(idx+2)+" <b>"+display_name+"</b>: PDF error: "+str(e))

            # Change report (all discrepancies + LOAs Generated sheet)
            try:
                report = generate_change_report(
                    changed_list, unchanged_list, new_list,
                    generated_changed, generated_unchanged)
                zf.writestr("Change_Report.xlsx", report)
            except Exception as e:
                warnings.append("Change report error: "+str(e))

        zip_buf.seek(0)
        token = secrets.token_hex(16)
        _zip_store[token] = zip_buf.read()

        self._json(200, {
            "changed_generated": len(generated_changed),
            "unchanged":         len(unchanged_list),
            "new_accts":         len(new_list),
            "token":             token,
            "warnings":          warnings,
        })

    def _handle_preview_accounts(self):
        """Return JSON {changed: [...], unchanged: [...]} for both pickers."""
        form = parse_multipart(self.headers, self.rfile)
        def fstr(k):
            v = form.get(k,"")
            return (v.decode("utf-8","replace") if isinstance(v, bytes) else v).strip()
        excel_bytes  = form.get("excel")
        hcg_bytes    = form.get("hcg")
        hcg_filename = fstr("hcg_filename")
        if not excel_bytes or not hcg_bytes:
            self._json(400, {"error":"Both files required."}); return
        try:
            df_xl = pd.read_excel(io.BytesIO(excel_bytes), dtype=str)
            df_xl.columns = [c.strip() for c in df_xl.columns]
        except Exception as e:
            self._json(400, {"error":"Could not read Excel: "+str(e)}); return
        try:
            df_hcg = read_hcg_report(hcg_bytes, hcg_filename or "report.csv")
        except Exception as e:
            self._json(400, {"error":"Could not read HCG report: "+str(e)}); return
        hcg_lookup = build_hcg_lookup(df_hcg)
        changed_names   = []
        unchanged_names = []
        for _, excel_row in df_xl.iterrows():
            excel_dict = dict(excel_row)
            if not clean(excel_row.get("Account Name","")): continue
            display_name = get_display_name(excel_dict)
            hcg_row = None
            for key in get_match_keys(excel_dict):
                if key in hcg_lookup:
                    hcg_row = hcg_lookup[key]; break
            if hcg_row is None: continue          # new account — not shown in pickers
            diffs, _ = compare_records(excel_dict, hcg_row)
            if diffs:
                changed_names.append(display_name)
            else:
                unchanged_names.append(display_name)
        self._json(200, {"changed": changed_names, "unchanged": unchanged_names})

    def _json(self, code, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type","application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers(); self.wfile.write(body)

# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    port = 5051
    server = HTTPServer(("127.0.0.1", port), LOAHandler)
    def _open():
        import time; time.sleep(1.2)
        webbrowser.open("http://127.0.0.1:"+str(port))
    threading.Thread(target=_open, daemon=True).start()
    print("\n  LOA Generator v2  ->  http://127.0.0.1:"+str(port))
    print("  Press Ctrl+C to stop.\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Stopped.")

if __name__ == "__main__":
    main()
